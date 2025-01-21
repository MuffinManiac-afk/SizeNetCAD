import numpy as np
import cv2
from PIL import Image
from scipy.cluster.vq import kmeans2
from matplotlib import pyplot as plt
import av
import socket
import struct
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
import pandas as pd
import math

# hl2ss之模組
from hl2ss_API import hl2ss_3dcv
import os
import sys

import torch
import warnings

def load_pose(file_path, sheet_name, number):
    wb = load_workbook(file_path, read_only=True)
    ws = wb[sheet_name]

    collecting = False
    rows_to_collect = 0
    data = []

    final_name = f"{sheet_name}_{number}"
    # final_name = f'{number}'
    for row in ws.iter_rows(values_only=True):
        if collecting:
            if rows_to_collect > 0:
                data.append(row[:4])
                rows_to_collect -= 1
                # data.append(row[:1])
            if rows_to_collect == 0:
                collecting = False
            continue

        if row[0] == final_name:
            rows_to_collect = 4
            collecting = True
        # elif row[0] == f"focal_length_{number}" or row[0] == f"principal_point_{number}":
        elif row[0] == f"fl_{number}" or row[0] == f"pp_{number}":

            rows_to_collect = 2
            collecting = True

    wb.close()
    return pd.DataFrame(data)

def Create_RGB_Intrinsics(intrinsics, focal_length, principal_point):
    intrinsics[0, 0] = focal_length[0]
    intrinsics[1, 1] = focal_length[1]
    intrinsics[0, 2] = principal_point[0]
    intrinsics[1, 2] = principal_point[1]
    return intrinsics

def depth_to_rgb_transit(calibration_lt, depth_images, xy1, RGB_focal_length, RGB_principal_point, RGB_pose, depth_pose):
    RGB_intrinsics = np.eye(3, 4, dtype=np.float32)
    R = np.array([[1, 0, 0, 0], [0, -1, 0, 0], [0, 0, -1, 0], [0, 0, 0, 1]], dtype=np.float32)

    ''' 座標轉換過程 '''
    RGB_intrinsics = Create_RGB_Intrinsics(RGB_intrinsics, RGB_focal_length, RGB_principal_point)
    depth_camera_points = xy1 * depth_images  # (288,320,3) # 深度感測器:像素座標系轉換至相機座標系
    homogenous_depth_camera_points = np.append(depth_camera_points, np.ones((288, 320, 1)),
                                                axis=2)  # (288,320,4):變齊次

    # reshape(-1, 4)：這部分將數據重新排列成一個形狀為 (n, 4) 的矩陣，其中 n 是根據數據的總長度自動計算出來的。-1 表示自動計算這個維度的大小。
    # .T：這是轉置操作，將矩陣的行和列互換。
    homogenous_depth_camera_points_reshaped = homogenous_depth_camera_points.reshape(-1, 4).T  # (4, 92160)

    # 深度感測器之相機座標系轉換至世界座標系所需的矩陣
    depth_camera_to_world_matrix = np.transpose(depth_pose) @ np.transpose(
        np.linalg.inv(calibration_lt.extrinsics))

    world_points = np.dot(depth_camera_to_world_matrix,
                            homogenous_depth_camera_points_reshaped)  # (4, 92160) : 世界座標

    # 世界座標轉換至RGB像素座標之矩陣 R:旋轉矩陣
    world_to_rgb_pixel_matrix = RGB_intrinsics @ R @ np.transpose(np.linalg.inv(RGB_pose))

    # 世界座標轉換至RGB像素座標
    depth_to_rgb_mapping_matrix = np.dot(world_to_rgb_pixel_matrix,
                                            world_points)  # (3, 92160)，但含有Zc，且為齊次(Zc[u,v,1])

    depth_to_rgb_mapping_matrix /= depth_to_rgb_mapping_matrix[-1, :]  # (3, 92160)，除上Zc，仍是齊次
    depth_to_rgb_mapping_matrix = depth_to_rgb_mapping_matrix[:2, :]  # (2, 92160)，消除齊次([u,v])
    depth_to_rgb_mapping_matrix = depth_to_rgb_mapping_matrix.T.reshape(288, 320, 2)  # 變成(288,320,2)
    # depth_to_rgb_mapping_matrix[10,30,0] => 深度影像像素座標(30,10) => rgb影像的x座標(水平)
    # depth_to_rgb_mapping_matrix[10,30,1] => 深度影像像素座標(30,10) => rgb影像的y座標(垂直)

    return depth_to_rgb_mapping_matrix

def check_valid_pixel_in_depth(depth_image, depth_x, depth_y, n):  # check 7*7內有無有效像素
    if depth_image[depth_y, depth_x] == 0:  # 無效像素
        if n == 0:  # x_min
            for row in range(depth_y + 3, depth_y - 4, -1):
                for col in range(depth_x - 3, depth_x + 4):
                    if (depth_y != row or depth_x != col) and depth_image[row, col] != 0:
                        return col, row
        elif n == 1:  # y_min
            for row in range(depth_y - 3, depth_y + 4):
                for col in range(depth_x - 3, depth_x + 4):
                    if (depth_y != row or depth_x != col) and depth_image[row, col] != 0:
                        return col, row
        elif n == 2:  # x_max
            for row in range(depth_y - 3, depth_y + 4):
                for col in range(depth_x + 3, depth_x - 4, -1):
                    if (depth_y != row or depth_x != col) and depth_image[row, col] != 0:
                        return col, row
        else:  # y_max
            for row in range(depth_y + 3, depth_y - 4, -1):
                for col in range(depth_x + 3, depth_x - 4, -1):
                    if (depth_y != row or depth_x != col) and depth_image[row, col] != 0:
                        return col, row

    return depth_x, depth_y

def find_nearest(depth_coords, target_coord):
    distances = np.sqrt((depth_coords[:, :, 0] - target_coord[0]) ** 2 + (depth_coords[:, :, 1] - target_coord[1]) ** 2)
    min_index = np.unravel_index(np.argmin(distances), distances.shape)
    return min_index

def filter(length_dict, preds, probs, num2class, tolerance_scope, estimated_target_max_length, top=100):  # 利用尺寸資訊過濾不符合的辨識結果
    preds_tmp = preds.detach().clone()[0][:top]  # 創建完全獨立的副本
    probs_tmp = probs.detach().clone()[0][:top]
    times = 0
    for i in range(0, top):
        P = length_dict[num2class[int(preds[0][i])]]
        if not ((P - tolerance_scope) <= estimated_target_max_length <= (P + tolerance_scope)):  # 尺寸不在此區間範圍的零件就擺至最後面
            preds_tmp = torch.cat([preds_tmp[preds_tmp != preds[0][i]], preds_tmp[preds_tmp == preds[0][i]]])
            probs_tmp = torch.cat([probs_tmp[probs_tmp != probs[0][i]], probs_tmp[probs_tmp == probs[0][i]]])

        new_preds = torch.unsqueeze(torch.tensor(preds_tmp).cpu(), dim=0)
        new_probs = torch.unsqueeze(torch.tensor(probs_tmp).cpu(), dim=0)

    return new_preds, new_probs

# def filter(length_dict, preds, probs, num2class, tolerance_scope, estimated_target_max_length): 
#     cmp_value = estimated_target_max_length
#     front, back = [], []
#     for i in range(0, len(preds[0])):
#         P = length_dict[num2class[int(preds[0][i])]]
#         if not ((P - tolerance_scope) <= cmp_value <= (P + tolerance_scope)):
#             back.append(i)
#         else:
#             front.append(i)

#     front_preds = torch.cat([preds[0][front], preds[0][back]])
#     front_probs = torch.cat([probs[0][front], probs[0][back]])
#     new_preds = torch.unsqueeze(front_preds.cuda(), dim=0)
#     new_probs = torch.unsqueeze(front_probs.cuda(), dim=0)
#     return new_preds, new_probs

def filter_v2(length_dict, preds, probs, num2class, tolerance_scope, estimated_target_max_length):
    preds_tmp = preds.detach().clone()
    probs_tmp = probs.detach().clone()
    times = 0
    for i in range(0, len(preds[0])):
        P = length_dict[num2class[int(preds[0][i])]]
        if not ((P - tolerance_scope) <= estimated_target_max_length <= (P + tolerance_scope)): 
            times +=1
            preds_tmp = torch.cat([preds_tmp[preds_tmp != preds[0][i]], preds_tmp[preds_tmp == preds[0][i]]])
            probs_tmp = torch.cat([probs_tmp[probs_tmp != probs[0][i]], probs_tmp[probs_tmp == probs[0][i]]])

        new_preds = torch.unsqueeze(torch.tensor(preds_tmp).cpu(), dim=0)
        new_probs = torch.unsqueeze(torch.tensor(probs_tmp).cpu(), dim=0)       
        if times == 15:
            break

    return new_preds, new_probs


def Calculate_Size(depth_image, lt_x, lt_y, rb_x, rb_y, depth_width, depth_height, calibration_lt):  # 估算物體大小

    min_x = lt_x
    min_y = lt_y
    max_x = rb_x
    max_y = rb_y

    tmp_depth_width = depth_width
    tmp_depth_height = depth_height
    
    if (max_x - min_x >= max_y - min_y):  # 代表寬的部分較長
        depth_width = max(tmp_depth_width, tmp_depth_height)
        depth_height = min(tmp_depth_width, tmp_depth_height)
    else:  # 代表長的部分較長
        depth_width = min(tmp_depth_width, tmp_depth_height)
        depth_height = max(tmp_depth_width, tmp_depth_height)
    
    if min_x > max_x:
        tmp = min_x
        min_x = max_x
        max_x = tmp
    if min_y > max_y:
        tmp = min_y
        min_y = max_y
        max_y = tmp

    # 找出指定物體在深度圖的範圍
    depth_PIL = Image.fromarray(depth_image)
    object_depth_image_PIL = depth_PIL.crop((min_x, min_y, max_x, max_y))
    object_depth_image = np.array(object_depth_image_PIL)

    # 找出中心點，以中心點的值當作基準
    center_x = int((min_x + max_x) // 2)
    center_y = int((min_y + max_y) // 2)
    distance = depth_image[center_y, center_x]
    # print(distance)

    sorted_object_depth_image = np.sort(np.array(object_depth_image))  # 由小到大排序
    non_zero = sorted_object_depth_image != 0

    if sorted_object_depth_image[non_zero].size == 0:
        # print("由於物體在深度影像中，沒辦法取得任何深度值，所以以整張深度影像中的最小值(不包含0)來代替距離。")
        distance = np.min(depth_image[depth_image != 0])
    elif (distance == 0):
        distance = np.median(sorted_object_depth_image[non_zero])
        # print("由於無法取得中心值，所以以所框區域中的中位數來代替。")
        # print(distance)
    elif (sorted_object_depth_image[non_zero].size > 0 and distance >= np.min(
            sorted_object_depth_image[non_zero] + 10)):
        non_zero_sorted_object_depth_image = sorted_object_depth_image[non_zero]
        upper_bound = np.min(non_zero_sorted_object_depth_image) + 10
        outliers = non_zero_sorted_object_depth_image >= upper_bound
        distance = np.median(non_zero_sorted_object_depth_image[~outliers])
        # print(distance)
    # else:
    #     # print(distance)

    # print('Depth Width: ', depth_width)
    # print('Depth Height: ', depth_height)
    # print('Calibration_lt_0_0 :' , calibration_lt.intrinsics[0, 0])
    # print('Calibration_lt_1_1 :' , calibration_lt.intrinsics[1, 1])
    e_width = (depth_width) * distance / calibration_lt.intrinsics[0, 0]
    e_height = (depth_height) * distance / calibration_lt.intrinsics[1, 1]
    return e_width, e_height


def estimate_length(depth_to_rgb_mapping_matrix, calibration_lt, image, depth_images,out):
    # 找出容納物體之矩形
    gray_img = cv2.cvtColor(image, cv2.COLOR_RGB2GRAY)  # 轉換為灰階影像
    gray_img = cv2.GaussianBlur(gray_img, (9, 9), 0)  # 高斯模糊
    edged = cv2.Canny(gray_img, 30, 60)  # Canny邊緣偵測
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (7, 7))
    edged = cv2.dilate(edged, kernel, iterations=2)  # 膨脹操作
    edged = cv2.erode(edged, kernel, iterations=2)  # 侵蝕操作

    contours, _ = cv2.findContours(edged, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)  # 找出影像中所有最外圍的輪廓 ; ok
    main_contour = max(contours, key=cv2.contourArea)  # 取這些輪廓中面積最大的 ; ok
    minRect = cv2.minAreaRect(main_contour)  # 根據這個輪廓，找到最小矩形，主要用於估算長度
    box = cv2.boxPoints(minRect)  #回傳矩形的四點座標
    left_top = min(box[:, 0]), min(box[:, 1])
    right_bottom = max(box[:, 0]), max(box[:, 1])
    cv2.drawContours(image, contours, -1, (0,255,0), 3)
    cv2.imwrite(out, image)
    # w = max(box_copy[:, 0]) - min(box_copy[:, 0])
    # h = max(box_copy[:, 1]) - min(box_copy[:, 1])

    # print(box)
    # print(left_top, right_bottom)   


    # RGB矩形頂點找到深度對應的地方
    lt_y, lt_x = find_nearest(depth_to_rgb_mapping_matrix, left_top)
    rb_y, rb_x = find_nearest(depth_to_rgb_mapping_matrix, right_bottom)
    box0_y, box0_x = find_nearest(depth_to_rgb_mapping_matrix, box[0])
    box1_y, box1_x = find_nearest(depth_to_rgb_mapping_matrix, box[1])
    box2_y, box2_x = find_nearest(depth_to_rgb_mapping_matrix, box[2])
    box3_y, box3_x = find_nearest(depth_to_rgb_mapping_matrix, box[3])

    # 深度影像中物體佔的像素寬度和長度(pixel)
    depth_width0 = np.sqrt((box0_x - box1_x) ** 2 + (box0_y - box1_y) ** 2)  # g1
    depth_height0 = np.sqrt((box1_x - box2_x) ** 2 + (box1_y - box2_y) ** 2)  # g2
    depth_width1 = np.sqrt((box2_x - box3_x) ** 2 + (box2_y - box3_y) ** 2)  # g1
    depth_height1 = np.sqrt((box3_x - box0_x) ** 2 + (box3_y - box0_y) ** 2)  # g2

    depth_obj_w = min(depth_width0, depth_width1)
    depth_obj_h = min(depth_height0, depth_height1)

    # 深度影像的處理
    depth_images = depth_images[:, :, 0]
    depth_images = depth_images * 100  # 公尺變公分

    # 估算物體大小(求最大長度,如無法求出,則返回-1)
    obj_width, obj_height = Calculate_Size(depth_images, lt_x, lt_y, rb_x, rb_y, depth_obj_w, depth_obj_h, calibration_lt)

    max_length = math.sqrt(obj_width**2 + obj_height**2)
    max_length = round(max_length, 6)
    max_length = max_length / 100
    
    # max_length = max(obj_width, obj_height) / 100

    return max_length


def masked_image(image):
    mask = np.zeros(image.shape[:2], dtype="uint8")
    top_left = (1000,186)
    bottom_right = (1710, 896)
    cv2.rectangle(mask, top_left, bottom_right, 255, -1)
    masked_image = cv2.bitwise_and(image, image, mask=mask)
    return masked_image

def get_length_dict(path):
    '''
        parse  log file to get real max length dict
        Args:
            path: log file path

        Returns: length_dict

        '''
    length_dict = {}
    with open(path, encoding='utf8') as f:
        lines = f.readlines()
        for i in range(0, len(lines)):
            if 'The size of ' in lines[i]:
                a=float(lines[i].split(':')[-1].split(',')[0].split('mm')[0].split(' ')[-1])/1000
                b=float(lines[i].split(':')[-1].split(',')[1].split('mm')[0].split(' ')[-1])/1000
                c=float(lines[i].split(':')[-1].split(',')[2].split('mm')[0].split(' ')[-1])/1000

                sorted_number = sorted([a, b, c], reverse=True)
                max_length = math.sqrt(sorted_number[0]**2 + sorted_number[1]**2) ###根據斜邊當作最長邊
                max_length = round(max_length, 6)
                # max_length = max(a,b,c)
                class_name = lines[i].split(' :')[0].split('The size of ')[-1].split('.iam')[0].split('.ipt')[0]
                if '_MIR' in class_name:
                    class_name=class_name.split('_MIR')[0]

                if class_name not in length_dict.keys():
                    length_dict[class_name] = max_length
    return length_dict